import os
from openpyxl import Workbook, load_workbook

dir_containing_files = r'C:\Users\patcharapol.y\Desktop\Projects\New folder\VS code\OCR\excels'
wb = Workbook()
for root, dir, filenames in os.walk(dir_containing_files):
    for file in filenames:
        if file.endswith('.xlsx'):
            file_name = file.split('.')[0]
            file_path = os.path.abspath(os.path.join(root, file))
            wb.create_sheet(file_name)
            dest_ws = wb[file_name]
            source_wb = load_workbook(file_path)
            source_sheet = source_wb.active
            for row in source_sheet.rows:
                for cell in row:
                    dest_ws[cell.coordinate] = cell.value
wb.save(os.path.join(r'C:\Users\patcharapol.y\Desktop\Projects\New folder\VS code\OCR\excels', 'all_raw_data.xlsx'))